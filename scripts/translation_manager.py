import os
import re
import polib
import itertools
import argparse
from typing import Dict, Tuple
from openpyxl import Workbook, load_workbook

_patterns = [
    ":c:func:`[^`]+`",
    ":c:type:`[^`]+`",
    ":c:macro:`[^`]+`",
    ":c:member:`[^`]+`",
    ":c:data:`[^`]+`",
    ":py:data:`[^`]+`",
    ":py:mod:`[^`]+`",
    ":func:`[^`]+`",
    ":mod:`[^`]+`",
    ":ref:`[^`]+`",
    ":class:`[^`]+`",
    ":pep:`[^`]+`",
    ":data:`[^`]+`",
    ":exc:`[^`]+`",
    ":term:`[^`]+`",
    ":meth:`[^`]+`",
    ":envvar:`[^`]+`",
    ":file:`[^`]+`",
    ":attr:`[^`]+`",
    ":const:`[^`]+`",
    ":issue:`[^`]+`",
    ":opcode:`[^`]+`",
    ":option:`[^`]+`",
    ":program:`[^`]+`",
    ":keyword:`[^`]+`",
    ":RFC:`[^`]+`",
    ":rfc:`[^`]+`",
    ":doc:`[^`]+`",
    ":source:`[^`]+`",
    ":manpage:`[^`]+`",
    ":mimetype:`[^`]+`",
    ":sup:`[^`]+`",
    ":kbd:`[^`]+`",
    ":const:`[^`]+`",
    "``[^`]+``",
    "`[^`]+`__",
    "`[^`]+`_",
    r"\*\*[^\*]+\*\*",  # bold text between **
    r"\*[^\*]+\*",  # italic text between *
]
_exps = [re.compile(e) for e in _patterns]


class Normalizer:
    @staticmethod
    def protect_sphinx_directives(s: str) -> Tuple[Dict[str, str], str]:
        """
        Replace Sphinx directives in the input string with placeholders.

        Parameters:
            s: The original string containing Sphinx directives.
        Returns:
            A tuple (placeholders, new_s) where 'placeholders' maps placeholders
            to the original directives and 'new_s' is the string with placeholders.
        """
        placeholders: Dict[str, str] = {}
        counter = itertools.count()

        def repl(match):
            ph = f"XASDF{next(counter):02}"
            placeholders[ph] = match.group(0)
            return ph

        combined_pattern = "|".join(f"({p})" for p in _patterns)
        combined_regex = re.compile(combined_pattern)
        new_s = combined_regex.sub(repl, s)
        return placeholders, new_s

    @staticmethod
    def undo_sphinx_directives_protection(
        placeholders: Dict[str, str], translated_text: str
    ) -> str:
        """
        Restore the original Sphinx directives in the translated text.
        """
        for ph, value in placeholders.items():
            translated_text = translated_text.replace(ph, value)
        return translated_text


class POExtractor:
    """Extracts unique strings from .po files in given directories and exports them to an Excel file."""

    def __init__(self, directories):
        self.directories = directories
        self.strings = {}

    def extract_strings_from_po(self, po_path: str) -> None:
        po = polib.pofile(po_path)
        for entry in po:
            if entry.msgid:
                self.strings[entry.msgid] = None

    def process_directories(self) -> None:
        for base_dir in self.directories:
            for root, _, files in os.walk(base_dir):
                for file in files:
                    if file.endswith(".po"):
                        po_path = os.path.join(root, file)
                        self.extract_strings_from_po(po_path)

    def generate_excel(self, output_excel: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Translations"
        ws.append(["Original", "Placeholders", "Temp Text", "Translation"])
        for original in self.strings:
            placeholders, temp_text = Normalizer.protect_sphinx_directives(original)
            ws.append([original, str(placeholders), temp_text, ""])
        wb.save(output_excel)
        print(f"Excel file saved to {output_excel}.")


class POUpdater:
    """Updates .po files with translations from an Excel file."""

    def __init__(self, directories, excel_file: str):
        self.directories = directories
        self.excel_file = excel_file
        self.translations = {}

    def load_translations(self) -> None:
        wb = load_workbook(self.excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                original = row[0]
                placeholders = eval(row[1])
                translated = row[3]
                translated = Normalizer.undo_sphinx_directives_protection(
                    placeholders, translated
                )
                self.translations[original] = translated

    def update_po_file(self, po_path: str) -> None:
        po = polib.pofile(po_path)
        updated = False
        for entry in po:
            if entry.msgid in self.translations:
                new_translation = self.translations[entry.msgid]
                if new_translation != entry.msgstr:
                    entry.msgstr = new_translation
                    if "fuzzy" not in entry.flags:
                        entry.flags.append("fuzzy")
                    updated = True
        if updated:
            po.save()
            print(f"Updated translations in {po_path}")

    def update_directories(self) -> None:
        self.load_translations()
        if not self.translations:
            print("No translations found in the Excel file.")
            return

        for base_dir in self.directories:
            for root, _, files in os.walk(base_dir):
                for file in files:
                    if file.endswith(".po"):
                        self.update_po_file(os.path.join(root, file))


def main():
    parser = argparse.ArgumentParser(
        description="Extract or update PO translations using an Excel file"
    )
    parser.add_argument(
        "mode",
        choices=["extract", "update"],
        help="Choose to extract from or update PO files",
    )
    parser.add_argument(
        "--dir",
        nargs="+",
        default=["./need_to_translate"],
        help="Directories containing PO files",
    )
    parser.add_argument(
        "--excel",
        type=str,
        default="translations.xlsx",
        help="Excel file to read from or write to",
    )
    args = parser.parse_args()

    if args.mode == "extract":
        extractor = POExtractor(args.dir)
        extractor.process_directories()
        extractor.generate_excel(args.excel)
    elif args.mode == "update":
        updater = POUpdater(args.dir, args.excel)
        updater.update_directories()


if __name__ == "__main__":
    main()
